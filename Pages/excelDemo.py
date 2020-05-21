import openpyxl

book = openpyxl.load_workbook("C:\\Users\\Srinath\\Desktop\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
#cell = sheet.cell(row=2,column=2)
#print(cell.value)

#sheet.cell(row=2,column=3).value = "HelloWorldie"
#print()
#book.save("C:\\Users\\Srinath\\Desktop\\PythonDemo.xlsx")

#Sheet row count
#print(sheet.max_row)

#Sheet column count
#print(sheet.max_column)
#
#print values in excel
for i in range(1,sheet.max_row):
    #if sheet.cell(row=i, column=1).value=="WomenPageExpandOptions":
    for j in range(1,sheet.max_column):
        Dict[sheet.cell(row=i+1,column=j).value]=sheet.cell(row=i+1,column=j+1).value

print(Dict)
#print(Dict["WomenLabelCheck"])
